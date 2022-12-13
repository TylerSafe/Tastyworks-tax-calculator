from openpyxl import load_workbook
import pathlib

PATH = pathlib.Path(__file__).parent.resolve()

try:
    wb = load_workbook(PATH + '\\History.csv')

    ws = wb['History']  
    date = ws['A']
    symbol = ws['M'] 
    desc = ws['F']
    value = ws['G']
    qty = ws['H']
    comm = ws['J']
    fees = ws['K']


    date = [ws['A'][x].value for x in range(len(ws['A']))]
    del date[0]

    print(date)

except:
    print("Error, no excel file found")